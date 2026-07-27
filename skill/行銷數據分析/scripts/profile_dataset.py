#!/usr/bin/env python3
"""
M1 資料剖析 —— 進場第一支腳本，把「這批資料到底是什麼」寫成白紙黑字。

為什麼需要它（沒有理由的腳本沒人會用）：
  M1 是整條產線唯一一個「錯了之後下游全部白做」的模組。粒度、主鍵、哪些值是假的
  —— 這三個判斷在 M1 做，而且只在 M1 做一次。漏掉的代價不是報告不夠嚴謹，是
  **每個數字都算對、但每個數字都在回答錯誤問題**。

  兩個實測案例就是這支腳本存在的全部理由（出處 references/04_資料體檢.md §4.1）：
    · Q2 哨兵值：課程檔 Step 2 的 `int` 欄含 100 個 9999，恰等於客戶數 100。
      不排除 → 平均購買間隔算成 199.46 天（真值 10.79，差 18.5 倍）→ 流失門檻訂成
      200 天 → 100 人只標出 3 人流失（真實 28 人），漏掉 25 人合計 NT$947,152。
      整條 pipeline 一路綠燈，不會報錯、不會有紅字。
    · Q3 樞紐雜訊列：Step 4 是 101 列 = 99 客戶 + Grand Total + (blank)。
      不剔除 → `Sum of Weight = 306,520`（全體加總）變成一位「顧客」，衝上任何
      排行榜第一名並拉高全體平均。

  所以本腳本做的不是「印一些描述統計」，是**把這兩類靜默污染在進 M2 之前攔下來**。

做了什麼（對應 04 §一 的步驟 ①③④）：
  ① 檔案盤點      magic bytes 驗真、解壓後大小、Excel 列上限截斷檢查
  ③ 逐欄剖析      型別／缺失率／唯一值數／五數摘要＋偏度峰度 CV／類別 top-10／
                   日期 min-max／疑似 ID 欄／疑似 helper 欄（規格見 04 §二）
  ④ 粒度與主鍵     單欄唯一 → 複合鍵 → 整列重複，決策樹見 04 §三
  ＋ 兩個特殊掃描  (a) 哨兵值（04 Q2）  (b) Excel 樞紐雜訊列（04 Q3）

  ② schema 契約比對交給 check_schema_contract.py；
  ⑤ 完整三桶品質關卡交給 check_data_quality.py（Q1/Q5/Q6/Q7/Q8/Q11/Q12/Q14/Q15
    需要契約、歷史載入或領域知識，不在剖析階段判定）。本腳本只回報自己看得見的。

輸出：
  · JSON  → stdout（ensure_ascii=False, indent=2），給下游腳本吃
  · 人眼報表 → stderr（分隔線格式），給人看
  · 落檔    → 開案與問題定義/欄位總表.csv（T1，欄位就是 04 §二 的 COLUMN_PROFILE_SCHEMA）
              開案與問題定義/資料剖析.json
              開案與問題定義/資料剖析報告.txt
  · 退出碼  → 0 通過｜1 有 error（擋住，不准進 M2）｜2 只有 warning｜3 腳本本身失敗
              與 setup_check.py、check_data_quality.py、verify_outputs 同一套，不准對調。

用法：
    python profile_dataset.py 2026Q3_信用卡                      # 掃專案的 原始資料/
    python profile_dataset.py 2026Q3_信用卡 a.parquet b.xlsx      # 指定檔案
    python profile_dataset.py 2026Q3_信用卡 --from-db             # 剖析專案 DuckDB 裡的表
    python profile_dataset.py 專案 x.parquet --report-to stdout   # 只看報表不吃 JSON
    python profile_dataset.py 專案 x.parquet --group-key 100      # 補一個外部已知的相異值數
    python profile_dataset.py 專案 x.xlsx --sheet Step2 --sheet Step4

實作判斷（reference 沒有明講、由本腳本自行決定的部分，一律標在這裡）：
  · unique_ratio 的分母用 n_total（總列數）而非 n_nonnull。04 §二 只寫「nunique/n」。
  · 「數值近似連續整數序列」定義為：整數欄且 nunique / (max - min + 1) ≥ 0.90。
  · 疑似 ID 欄的「型別為整數或字串」放寬到「整數值的 float」—— 否則含 1 個 NaN 的
    整數欄會被 pandas 轉成 float64 而漏判（課程檔的 `Unnamed: 10` 正是這種）。
  · 哨兵值的 group-by 鍵候選取自**同一個檔案的所有表**，不限本表；命中時會標明是
    哪張表的哪一欄。理由：樞紐/中間計算表的哨兵常常對應到主檔的實體數。
  · 「極接近」預設為相對誤差 ≤ 2%（--sentinel-tolerance 可調），且要求命中的鍵
    相異值數 ≥ 3、哨兵出現次數 ≥ 3，避免小數字亂命中。
  · Excel 列上限截斷（1,048,575／1,048,576 列）判為 error。04 §一 把它列為檔案盤點
    必查項但未分桶；判 error 的理由是它會靜默丟掉整批列（實測遺失約 248,100 列），
    與 Q2 同一類「不報錯但結論全錯」的污染。
  · practical_use（04 §二 的招牌欄）由欄名與型別自動推定，一律標記「需人工確認」。
    自動推定不算數，04 §二 說得很清楚：填不出來代表這欄還沒被想清楚。
  · 「沒有檔案可剖析」「指定的檔案不存在」判 error（退出碼 1）而非 3。退出碼 3 依
    04 §四 只代表「檢查腳本本身失敗」，留給最外層的例外處理，不拿來表示使用者失誤。
  · --no-pivot-clean 只停止「剔除」，不停止「偵測」：命中仍然是 error，並額外註明
    本次所有統計量都含那幾列。實測對照（Step 4 的 Sum of Weight）：
      剔除 max = 50,403、mean = 3,096.16｜保留 max = 306,520、mean = 6,130.40
    —— Grand Total 一列就讓平均翻倍、最大值變成 6 倍，這是 04 §三 說的「衝上榜首」。
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Sequence

import numpy as np
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent))
from paths import project_dir  # noqa: E402

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

VERSION = "1.0"

# ── 04 §4.1 Q2：哨兵值候選集，不准擅自增刪（要改先改 reference）────────
SENTINEL_NUMERIC: tuple[float, ...] = (-1, -9, -99, -999, 0, 9999, 99999, 999999)
SENTINEL_DATES: tuple[str, ...] = ("1900-01-01", "1970-01-01", "2099-12-31", "9999-12-31")
# 0 太常見，單獨列出：只在「命中某個 group-by 鍵的相異值數」時才報，不報未命中的
_SOFT_SENTINELS = {0}

# ── 04 §4.1 Q3：Excel 樞紐雜訊列的列標籤 ────────────────────────────
PIVOT_NOISE_LABELS = ("grand total", "總計", "合計", "(blank)", "(空白)")

# ── 04 §二：缺失率門檻 ──────────────────────────────────────────────
NULL_WARN = 0.05
NULL_ERROR = 0.40
NULL_DROP = 0.80

# ── 04 §二：疑似 ID 欄 / helper 欄 ──────────────────────────────────
ID_RATIO = 0.95
ID_NAME_RE = re.compile(r"(id|編號|序號|代碼|代号|key|no\b|_no$)", re.IGNORECASE)
SUBJECT_NAME_RE = re.compile(r"(客戶|顧客|會員|customer|member|user|cust)", re.IGNORECASE)
HELPER_NAME_RE = re.compile(r"^Unnamed:\s*\d+$")
SEQUENCE_RATIO = 0.90

# ── 04 §二：形狀量數門檻 ────────────────────────────────────────────
SKEW_ABS = 1.0
CV_HIGH = 1.0

# ── 04 §4.3：info 桶門檻 ────────────────────────────────────────────
NEAR_CONSTANT = 0.95
HIGH_CARDINALITY = 50

# ── 04 §一：Excel 列上限（被截斷的指紋）──────────────────────────────
EXCEL_ROW_LIMITS = (1_048_575, 1_048_576)

# ── 04 §4.2 Q16：金額欄與幣別欄的欄名樣式 ───────────────────────────
MONEY_NAME_RE = re.compile(r"(金額|價格|營收|銷售額|消費|spend|amount|price|revenue|sales|cost)", re.IGNORECASE)
CURRENCY_NAME_RE = re.compile(r"(幣別|幣種|currency|ccy)", re.IGNORECASE)

# ── 04 §4.2 Q13：類別編碼前綴 ───────────────────────────────────────
CODE_PREFIX_RE = re.compile(r"^([0-9]{1,3})([_.\-])")

# ── 04 §4.2 Q10：全形區與多餘空白 ───────────────────────────────────
FULLWIDTH_RE = re.compile(r"[！-～　]")

# ── 04 §二：分群輸入白名單（見 18-E2，只能是行為指標）────────────────
SEGMENT_NAME_RE = re.compile(r"^(r|f|m|rfm.*|cai|cri|ln_[fm]|factor.*|f\d+)$", re.IGNORECASE)
# ── 04 §二：人口統計變數只能 profile_only，不准進分群（見 18-E2）──────
DEMOGRAPHIC_NAME_RE = re.compile(
    r"(年齡|age|性別|gender|sex|生日|birth|居住地|地區|城市|region|city|zip|"
    r"教育|education|婚姻|marital|職業|occupation|job|所得|收入|income|學歷)",
    re.IGNORECASE)


# ══════════════════════════════════════════════════════════════════
# 三桶
# ══════════════════════════════════════════════════════════════════
@dataclass
class Finding:
    """一條發現。訊息一律「事實 — 該怎麼辦」兩段式，不要只報錯不給出路。"""
    severity: str          # error / warning / info
    rule: str              # Q2 / Q3 / 缺失率 / …
    where: str             # 檔案::表::欄
    fact: str
    action: str

    def to_dict(self) -> dict[str, str]:
        return {"嚴重度": self.severity, "規則": self.rule, "位置": self.where,
                "事實": self.fact, "該怎麼辦": self.action}


@dataclass
class Buckets:
    error: list[Finding] = field(default_factory=list)
    warning: list[Finding] = field(default_factory=list)
    info: list[Finding] = field(default_factory=list)

    def add(self, severity: str, rule: str, where: str, fact: str, action: str) -> None:
        getattr(self, severity).append(Finding(severity, rule, where, fact, action))

    def gate(self) -> int:
        if self.error:
            return 1
        if self.warning:
            return 2
        return 0


# ══════════════════════════════════════════════════════════════════
# 小工具
# ══════════════════════════════════════════════════════════════════
def _jsonable(o: Any) -> Any:
    """把 numpy / pandas / Path 轉成 json 吞得下的型別。"""
    if o is None or isinstance(o, (str, bool)):
        return o
    if isinstance(o, (np.integer,)):
        return int(o)
    if isinstance(o, (np.floating, float)):
        v = float(o)
        return None if (np.isnan(v) or np.isinf(v)) else v
    if isinstance(o, (int,)):
        return o
    if isinstance(o, (np.bool_,)):
        return bool(o)
    if isinstance(o, (pd.Timestamp, datetime, date)):
        return o.isoformat()
    if isinstance(o, pd.Timedelta):
        return str(o)
    if isinstance(o, Path):
        return str(o)
    if isinstance(o, dict):
        return {str(k): _jsonable(v) for k, v in o.items()}
    if isinstance(o, (list, tuple, set, np.ndarray, pd.Index)):
        return [_jsonable(v) for v in o]
    if o is pd.NaT or (isinstance(o, float) and np.isnan(o)):
        return None
    return str(o)


def _fmt(v: Any, nd: int = 4) -> str:
    if v is None:
        return "—"
    if isinstance(v, float):
        if np.isnan(v) or np.isinf(v):
            return "—"
        if abs(v) >= 1000:
            return f"{v:,.2f}"
        return f"{v:.{nd}g}"
    return str(v)


def _pct(v: float | None) -> str:
    return "—" if v is None else f"{v * 100:.2f}%"


def _w(s: str, width: int) -> str:
    """依顯示寬度切齊（中文算兩格）。報表對齊用。"""
    out, used = [], 0
    for ch in str(s):
        cw = 2 if ord(ch) > 0x1100 and not 0xFF61 <= ord(ch) <= 0xFF9F else 1
        if used + cw > width:
            break
        out.append(ch)
        used += cw
    return "".join(out) + " " * max(0, width - used)


def _norm_for_compare(s: pd.Series) -> pd.Series:
    """把欄位正規化成可以直接比對的字串。

    整數值的 float 要先變回整數，否則 89.0 與 89 會被判成不同 —— 課程檔的
    `Unnamed: 10`（含 1 個 NaN 被轉成 float64）與 `客戶ID`（int64）正是這種，
    比不出來就抓不到「helper 欄其實是主鍵複本」這個 18-E10 的實例。
    """
    if _is_integral(s):
        return s.astype("float64").astype("Int64").astype("string")
    return s.astype("string")


def _is_integral(s: pd.Series) -> bool:
    """整數值的 float 也算整數欄 —— 含 1 個 NaN 就被轉 float64 是 pandas 的常態。"""
    v = s.dropna()
    if v.empty:
        return False
    if pd.api.types.is_integer_dtype(v):
        return True
    if not pd.api.types.is_float_dtype(v):
        return False
    arr = v.to_numpy(dtype="float64", copy=False)
    return bool(np.isfinite(arr).all() and np.all(np.mod(arr, 1) == 0))


# ══════════════════════════════════════════════════════════════════
# ① 檔案盤點
# ══════════════════════════════════════════════════════════════════
def sniff_format(path: Path) -> tuple[str, str]:
    """用 magic bytes 判真格式。回傳 (真格式, 說明)。04 §一：以 magic bytes 為準。"""
    try:
        head = path.open("rb").read(8)
    except OSError as e:
        return "unknown", f"讀不到檔頭：{e}"
    if head[:4] == b"PAR1":
        return "parquet", "PAR1"
    if head[:4] == b"PK\x03\x04":
        try:
            with zipfile.ZipFile(path) as z:
                names = z.namelist()
            if "[Content_Types].xml" in names:
                return "xlsx", "504b0304 + [Content_Types].xml（真 OOXML）"
            return "zip", "504b0304（zip，但不是 OOXML）"
        except zipfile.BadZipFile:
            return "zip", "504b0304 但 zip 解析失敗"
    if head[:2] == b"\xd0\xcf":
        return "xls", "D0CF11E0（舊版 OLE2）"
    if head[:8] == b"SQLite f" or head[:4] == b"DUCK":
        return "duckdb", head[:8].decode("latin-1", "replace")
    return "text", "無二進位指紋，視為文字（csv/tsv/json）"


def xlsx_uncompressed(path: Path) -> dict[str, int]:
    """回傳 zip 成員解壓後大小。04 §一：能不能整份載入要看這個，不是看檔案大小。"""
    out: dict[str, int] = {}
    try:
        with zipfile.ZipFile(path) as z:
            for i in z.infolist():
                out[i.filename] = i.file_size
    except (zipfile.BadZipFile, OSError):
        pass
    return out


def inventory(path: Path, bk: Buckets, big_mb: int) -> dict[str, Any]:
    real, why = sniff_format(path)
    claimed = path.suffix.lower().lstrip(".")
    size = path.stat().st_size if path.exists() else 0
    inv: dict[str, Any] = {
        "路徑": str(path), "檔名": path.name,
        "位元組": size, "MB": round(size / 1024 / 1024, 2),
        "宣稱格式": claimed, "實際格式": real, "magic": why,
        "副檔名詐騙": False, "解壓後大小MB": None, "可整份載入": True,
    }
    equiv = {"xlsx": {"xlsx", "xlsm", "xltx"}, "text": {"csv", "tsv", "txt", "json"},
             "parquet": {"parquet", "pq"}, "xls": {"xls"}, "duckdb": {"duckdb", "db"}}
    if claimed and claimed not in equiv.get(real, {real}):
        inv["副檔名詐騙"] = True
        bk.add("warning", "副檔名詐騙", path.name,
               f"副檔名是 .{claimed}，magic bytes 說它是 {real}（{why}）",
               f"以 magic bytes 為準，本腳本已改用 {real} 讀法。"
               f"請確認上游匯出流程是否用錯副檔名")

    if real == "xlsx":
        members = xlsx_uncompressed(path)
        total = sum(members.values())
        inv["解壓後大小MB"] = round(total / 1024 / 1024, 2)
        biggest = sorted(members.items(), key=lambda kv: -kv[1])[:3]
        inv["最大成員"] = [{"成員": k, "MB": round(v / 1024 / 1024, 2)} for k, v in biggest]
        if total / 1024 / 1024 > big_mb:
            inv["可整份載入"] = False
            bk.add("warning", "大檔", path.name,
                   f"解壓後共 {inv['解壓後大小MB']} MB（超過 --big-mb {big_mb}），"
                   f"最大成員 {biggest[0][0]} {round(biggest[0][1]/1024/1024, 2)} MB",
                   "pandas.read_excel 峰值可達解壓後大小的 8–10 倍，本腳本改走 "
                   "openpyxl read_only 串流；要完整剖析請先轉成 parquet")
    return inv


# ══════════════════════════════════════════════════════════════════
# 讀表
# ══════════════════════════════════════════════════════════════════
def _read_csv(path: Path, bk: Buckets) -> pd.DataFrame:
    last: Exception | None = None
    for enc in ("utf-8-sig", "utf-8", "cp950", "big5", "latin-1"):
        try:
            df = pd.read_csv(path, encoding=enc, low_memory=False)
            if enc not in ("utf-8-sig", "utf-8"):
                bk.add("warning", "編碼", path.name,
                       f"UTF-8 解不開，改用 {enc} 讀成功",
                       "台灣資料常見 Big5/CP950。轉檔進 staging 時一律改存 UTF-8，"
                       "DuckDB 端見 db.py 的 encodings 擴充")
            return df
        except (UnicodeDecodeError, LookupError) as e:
            last = e
    raise RuntimeError(f"CSV 編碼全部試過都失敗：{last}")


def _read_xlsx_stream(path: Path, sheet: str, max_rows: int) -> pd.DataFrame:
    """大檔走串流，不用 pandas.read_excel（04 §一）。

    一律餵 binary handle 而不是路徑：openpyxl 與 pandas 都會照副檔名擋人，
    而「副檔名詐騙」正是 04 §一 的必查項之一（實測有 227 MB 的 xlsx 叫 .csv.xlsx）。
    """
    import openpyxl
    with path.open("rb") as fh:
        wb = openpyxl.load_workbook(fh, read_only=True, data_only=True)
        ws = wb[sheet]
        rows: list[tuple] = []
        header: list[str] | None = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header = [str(c) if c is not None else f"Unnamed: {j}"
                          for j, c in enumerate(row)]
                continue
            rows.append(row)
            if len(rows) >= max_rows:
                break
        wb.close()
    return pd.DataFrame(rows, columns=header)


def read_tables(path: Path, inv: dict[str, Any], bk: Buckets,
                sheets: Sequence[str] | None, max_rows: int) -> dict[str, pd.DataFrame]:
    real = inv["實際格式"]
    if real == "parquet":
        return {path.stem: pd.read_parquet(path)}
    if real == "text":
        return {path.stem: _read_csv(path, bk)}
    if real in ("xlsx", "xls"):
        import openpyxl
        if real == "xlsx":
            with path.open("rb") as fh:
                wb = openpyxl.load_workbook(fh, read_only=True, data_only=True)
                names = list(wb.sheetnames)
                wb.close()
        else:
            names = list(pd.ExcelFile(path).sheet_names)
        want = [s for s in names if (not sheets or s in sheets)]
        if sheets:
            for s in sheets:
                if s not in names:
                    bk.add("warning", "指定分頁不存在", f"{path.name}::{s}",
                           f"--sheet {s} 在活頁簿裡找不到（實有：{'、'.join(names)}）",
                           "確認分頁名稱，或拿掉 --sheet 讓腳本掃全部")
        out: dict[str, pd.DataFrame] = {}
        for s in want:
            if inv["可整份載入"]:
                # engine 要明講、且餵 handle：檔案被改名成 .csv 時
                # pandas 與 openpyxl 都會照副檔名挑錯引擎或直接拒收
                if real == "xlsx":
                    with path.open("rb") as fh:
                        out[s] = pd.read_excel(fh, sheet_name=s, engine="openpyxl")
                else:
                    out[s] = pd.read_excel(path, sheet_name=s)
            else:
                out[s] = _read_xlsx_stream(path, s, max_rows)
                bk.add("info", "串流讀取", f"{path.name}::{s}",
                       f"大檔走 openpyxl 串流，本次只讀前 {max_rows:,} 列",
                       "統計量是抽樣值不是母體值，報告不可直接引用。"
                       "要完整剖析請先用 DuckDB 轉成 parquet")
        return out
    raise RuntimeError(f"不支援的格式：{real}（{inv['magic']}）")


# ══════════════════════════════════════════════════════════════════
# 特殊掃描 (b)：Excel 樞紐雜訊列（04 §4.1 Q3）
# ══════════════════════════════════════════════════════════════════
def strip_pivot_noise(df: pd.DataFrame, extra: Sequence[str],
                      enabled: bool = True) -> tuple[pd.DataFrame, dict[str, Any]]:
    """剔除列標籤為 Grand Total／總計／合計／(blank)／(空白) 的列。

    只看**首欄**（04 Q3：首欄值命中）。剔除的列會完整記錄下來，不是默默丟掉。
    enabled=False（--no-pivot-clean）時只偵測不剔除 —— 有些表的「合計」真的是
    一個業務類別，這時要人來判斷，但仍然要讓他知道命中了什麼。
    """
    info: dict[str, Any] = {"已剔除列數": 0, "命中標籤": [], "剔除的列": [],
                            "偵測到但未剔除": 0}
    if df.empty or df.shape[1] == 0:
        return df, info
    labels = {s.strip().casefold() for s in list(PIVOT_NOISE_LABELS) + list(extra)}
    first = df.columns[0]
    col = df[first]
    if not (pd.api.types.is_object_dtype(col) or isinstance(col.dtype, pd.CategoricalDtype)):
        return df, info
    norm = col.astype("string").str.strip().str.casefold()
    hit = norm.isin(labels).fillna(False)
    if not hit.any():
        return df, info
    dropped = df.loc[hit]
    info["命中標籤"] = [str(v) for v in col[hit].unique()]
    info["剔除的列"] = _jsonable(dropped.head(10).to_dict(orient="records"))
    info["首欄"] = str(first)
    if not enabled:
        info["偵測到但未剔除"] = int(hit.sum())
        return df, info
    info["已剔除列數"] = int(hit.sum())
    return df.loc[~hit].reset_index(drop=True), info


# ══════════════════════════════════════════════════════════════════
# ③ 逐欄剖析（04 §二）
# ══════════════════════════════════════════════════════════════════
def _infer_dtype(s: pd.Series) -> str:
    if pd.api.types.is_datetime64_any_dtype(s):
        return "datetime64[ns]"
    if pd.api.types.is_bool_dtype(s):
        return "bool"
    if pd.api.types.is_integer_dtype(s):
        return "int64"
    if pd.api.types.is_float_dtype(s):
        return "int64(float 儲存)" if _is_integral(s) else "float64"
    return "object"


def _looks_sequence(s: pd.Series) -> bool:
    v = s.dropna()
    if v.empty or not _is_integral(v):
        return False
    arr = v.to_numpy(dtype="float64", copy=False)
    span = arr.max() - arr.min() + 1
    return span > 0 and (len(np.unique(arr)) / span) >= SEQUENCE_RATIO


def _practical_use(name: str, s: pd.Series, is_id: bool, is_helper: bool,
                   is_pk_part: bool) -> str:
    """04 §二 的七種用途標籤。自動推定，一律要人工確認。"""
    if is_helper:
        return "helper"
    if is_id or ID_NAME_RE.search(name):
        if SUBJECT_NAME_RE.search(name):
            return "subject_key"
        return "subject_key" if is_pk_part and is_id else "fk"
    # 人口統計變數優先於一切 —— 18-E2：它們只能出現在卡方那一章，不准進分群
    if DEMOGRAPHIC_NAME_RE.search(name):
        return "profile_only"
    # segment_input 必須是數值型的行為指標；名字叫 CAI 但值是 Present/Missing 的
    # 旗標欄不算（課程檔的客戶檔就有這一欄）
    if SEGMENT_NAME_RE.match(name.strip()) and pd.api.types.is_numeric_dtype(s):
        return "segment_input"
    if pd.api.types.is_datetime64_any_dtype(s):
        return "time_axis"
    if MONEY_NAME_RE.search(name) or (
        pd.api.types.is_numeric_dtype(s) and not pd.api.types.is_bool_dtype(s)
    ):
        return "measure"
    return "profile_only"


def profile_column(df: pd.DataFrame, col: str, ordinal: int, source: str,
                   table: str, top_n: int) -> dict[str, Any]:
    s = df[col]
    n = len(df)
    n_nonnull = int(s.notna().sum())
    try:
        n_unique = int(s.nunique(dropna=True))
    except TypeError:                        # 不可雜湊（list 欄）
        n_unique = int(s.astype(str).nunique())
    null_rate = 1.0 - (n_nonnull / n) if n else 0.0
    ratio = (n_unique / n) if n else 0.0

    dtype_inf = _infer_dtype(s)
    is_helper = bool(HELPER_NAME_RE.match(str(col))) or n_nonnull == 0
    is_intlike = _is_integral(s) or pd.api.types.is_string_dtype(s) or \
        pd.api.types.is_object_dtype(s)
    is_id = bool(
        ratio >= ID_RATIO and n_nonnull > 0 and is_intlike
        and not pd.api.types.is_datetime64_any_dtype(s)
        and (ID_NAME_RE.search(str(col)) or _looks_sequence(s))
    )

    rec: dict[str, Any] = {
        "source": source, "table": table, "column": str(col), "ordinal": ordinal,
        "dtype_raw": str(s.dtype), "dtype_inferred": dtype_inf,
        "n_total": n, "n_nonnull": n_nonnull, "null_rate": round(null_rate, 6),
        "n_unique": n_unique, "unique_ratio": round(ratio, 6),
        "min": None, "q25": None, "median": None, "q75": None, "max": None,
        "mean": None, "std": None, "p95": None, "p99": None,
        "skewness": None, "kurtosis": None, "cv": None,
        "top10_values": None, "top10_shares": None, "tail_share": None,
        "date_min": None, "date_max": None,
        "n_distinct_dates": None, "n_missing_dates": None,
        "suspected_id": is_id, "helper_column": is_helper,
        "suspected_sentinel": False,        # 由 scan_sentinels 回填
        "practical_use": "", "note": "",
    }

    notes: list[str] = []

    # ── 數值欄：五數摘要 + 形狀量數 ──────────────────────────────
    if pd.api.types.is_numeric_dtype(s) and not pd.api.types.is_bool_dtype(s):
        v = pd.to_numeric(s, errors="coerce").dropna().to_numpy(dtype="float64")
        v = v[np.isfinite(v)]
        if v.size:
            rec.update(
                min=float(v.min()), max=float(v.max()), mean=float(v.mean()),
                q25=float(np.percentile(v, 25)), median=float(np.percentile(v, 50)),
                q75=float(np.percentile(v, 75)),
                p95=float(np.percentile(v, 95)), p99=float(np.percentile(v, 99)),
                std=float(v.std(ddof=1)) if v.size > 1 else 0.0,
            )
            try:
                from scipy import stats as _st
                if v.size >= 3:
                    rec["skewness"] = float(_st.skew(v, bias=False))
                if v.size >= 4:
                    rec["kurtosis"] = float(_st.kurtosis(v, bias=False))
            except ImportError:
                notes.append("scipy 缺席，偏度峰度未算")
            # 04 §二：含負值或含 0 一律不報 CV
            if rec["min"] is not None and rec["min"] > 0 and rec["mean"]:
                rec["cv"] = float(rec["std"] / rec["mean"])
            else:
                notes.append("含 0 或負值，依 04 §二 不報 CV")

    # ── 日期欄 ────────────────────────────────────────────────
    if pd.api.types.is_datetime64_any_dtype(s):
        v = s.dropna()
        if not v.empty:
            d = v.dt.normalize()
            rec["date_min"] = d.min().date().isoformat()
            rec["date_max"] = d.max().date().isoformat()
            nd = int(d.nunique())
            span = int((d.max() - d.min()).days) + 1
            rec["n_distinct_dates"] = nd
            rec["n_missing_dates"] = max(0, span - nd)

    # ── 類別 top-N ────────────────────────────────────────────
    wants_top = (
        pd.api.types.is_object_dtype(s) or isinstance(s.dtype, pd.CategoricalDtype)
        or pd.api.types.is_bool_dtype(s) or pd.api.types.is_string_dtype(s)
        or (n_unique <= 20 and not pd.api.types.is_datetime64_any_dtype(s))
    )
    # 幾乎每列一個值的欄（鍵欄）列 top-N 沒有意義，只會刷版面
    if ratio >= ID_RATIO and n_unique > top_n:
        wants_top = False
        notes.append("接近逐列唯一，不列 top-N")
    if wants_top and n_nonnull:
        try:
            vc = s.value_counts(dropna=True)
        except TypeError:
            vc = s.astype(str).value_counts(dropna=True)
        head = vc.head(top_n)
        rec["top10_values"] = [str(i) for i in head.index]
        rec["top10_shares"] = [round(float(c) / n_nonnull, 6) for c in head.to_numpy()]
        rec["tail_share"] = round(float(vc.iloc[top_n:].sum()) / n_nonnull, 6) \
            if len(vc) > top_n else 0.0

    rec["practical_use"] = _practical_use(str(col), s, is_id, is_helper, False)
    rec["note"] = "；".join(notes)
    return rec


# ══════════════════════════════════════════════════════════════════
# ④ 粒度與主鍵（04 §三 決策樹）
# ══════════════════════════════════════════════════════════════════
def _dup_count(df: pd.DataFrame, cols: Sequence[str]) -> int:
    try:
        return int(df.duplicated(subset=list(cols)).sum())
    except TypeError:
        return int(df[list(cols)].astype(str).duplicated().sum())


def infer_grain(df: pd.DataFrame, cols: list[dict[str, Any]],
                max_key_cols: int) -> dict[str, Any]:
    n = len(df)
    out: dict[str, Any] = {
        "列數": n, "主鍵": None, "主鍵種類": None, "替代鍵": [],
        "粒度描述": None, "整列重複組數": 0, "整列重複列數": 0,
        "最接近的複合鍵": None, "最接近複合鍵重複率": None,
    }

    # 整列完全重複 —— 04 §三：只有「所有欄位皆同」才算真重複，不准放寬
    try:
        dup_mask = df.duplicated(keep=False)
    except TypeError:
        dup_mask = df.astype(str).duplicated(keep=False)
    n_dup_rows = int(dup_mask.sum())
    out["整列重複列數"] = n_dup_rows
    if n_dup_rows:
        try:
            out["整列重複組數"] = int(df.duplicated(keep="first").sum())
        except TypeError:
            out["整列重複組數"] = int(df.astype(str).duplicated(keep="first").sum())

    if n == 0:
        out["粒度描述"] = "空表，無法判定粒度"
        return out

    # 第一層：單欄唯一。float 欄不當鍵 —— 浮點值相等是巧合不是設計
    singles = [c for c in cols
               if c["n_unique"] == n and c["n_nonnull"] == n and not c["helper_column"]
               and c["dtype_inferred"] != "float64"]
    if singles:
        def rank(c: dict[str, Any]) -> tuple[int, int, int]:
            return (0 if c["suspected_id"] else 1,
                    0 if ID_NAME_RE.search(c["column"]) else 1,
                    c["ordinal"])
        singles.sort(key=rank)
        pk = singles[0]["column"]
        out["主鍵"] = [pk]
        out["主鍵種類"] = "單欄"
        out["替代鍵"] = [c["column"] for c in singles[1:]]
        out["粒度描述"] = f"一列 = 一個「{pk}」"
        return out

    # 第二層：複合鍵。候選欄依 04 §三 建議的組合語意排序
    def cand_rank(c: dict[str, Any]) -> tuple[int, int]:
        name = c["column"]
        if SUBJECT_NAME_RE.search(name) and ID_NAME_RE.search(name):
            g = 0
        elif c["dtype_inferred"] == "datetime64[ns]":
            g = 1
        elif ID_NAME_RE.search(name) or c["suspected_id"]:
            g = 2
        elif c["n_unique"] <= max(50, n // 10):
            g = 3
        else:
            g = 4
        return (g, c["ordinal"])

    cands = [c for c in cols
             if not c["helper_column"] and c["n_unique"] > 1 and c["n_nonnull"] == n
             and c["dtype_inferred"] != "float64"]
    cands.sort(key=cand_rank)
    cands = cands[:max_key_cols]
    names = [c["column"] for c in cands]

    best: tuple[int, tuple[str, ...]] | None = None
    from itertools import combinations
    for k in (2, 3):
        if k > len(names):
            break
        for combo in combinations(names, k):
            d = _dup_count(df, combo)
            if best is None or d < best[0]:
                best = (d, combo)
            if d == 0:
                out["主鍵"] = list(combo)
                out["主鍵種類"] = f"{k} 欄複合鍵"
                out["粒度描述"] = "一列 = 一個「" + " × ".join(combo) + "」組合"
                return out
        if best and best[0] == 0:
            break

    if best:
        out["最接近的複合鍵"] = list(best[1])
        out["最接近複合鍵重複率"] = round(best[0] / n, 6)
    out["粒度描述"] = (
        "本表無主鍵 —— 單欄與 2/3 欄複合鍵都不唯一。"
        + (f"最接近的是「{' × '.join(best[1])}」，仍有 {best[0]:,} 列重複"
           f"（{best[0] / n:.2%}）" if best else "")
    )
    return out


# ══════════════════════════════════════════════════════════════════
# 特殊掃描 (a)：哨兵值（04 §4.1 Q2）
# ══════════════════════════════════════════════════════════════════
@dataclass
class GroupKey:
    table: str
    column: str
    n_distinct: int


def collect_group_keys(table: str, cols: list[dict[str, Any]]) -> list[GroupKey]:
    """挑出可以當 group-by 鍵的欄。哨兵值就是拿出現次數去比它們的相異值數。

    課程檔的關鍵案例：`Step 2` 的 `客戶ID` 相異值 100（不是唯一鍵，unique_ratio
    只有 0.019），所以候選池不能只收 suspected_id，否則 9999×100 這條永遠抓不到。
    """
    out: list[GroupKey] = []
    for c in cols:
        if c["helper_column"] or c["n_unique"] < 3:
            continue
        if c["dtype_inferred"] in ("float64",):
            continue
        looks_key = (
            bool(ID_NAME_RE.search(c["column"]))
            or c["suspected_id"]
            or (c["unique_ratio"] <= 0.5 and c["dtype_inferred"] != "datetime64[ns]")
        )
        if looks_key:
            out.append(GroupKey(table, c["column"], c["n_unique"]))
    return out


def scan_sentinels(df: pd.DataFrame, cols: list[dict[str, Any]], table: str,
                   keys: list[GroupKey], tol: float,
                   min_count: int) -> list[dict[str, Any]]:
    hits: list[dict[str, Any]] = []
    date_cands = [pd.Timestamp(d) for d in SENTINEL_DATES]

    for c in cols:
        col = c["column"]
        s = df[col]
        if c["n_nonnull"] == 0:
            continue

        found: list[tuple[Any, int]] = []
        if pd.api.types.is_datetime64_any_dtype(s):
            d = s.dropna().dt.normalize()
            for cand in date_cands:
                cnt = int((d == cand).sum())
                if cnt:
                    found.append((cand.date().isoformat(), cnt))
        elif pd.api.types.is_numeric_dtype(s) and not pd.api.types.is_bool_dtype(s):
            v = pd.to_numeric(s, errors="coerce")
            for cand in SENTINEL_NUMERIC:
                cnt = int((v == cand).sum())
                if cnt:
                    found.append((cand, cnt))
        elif pd.api.types.is_object_dtype(s) or pd.api.types.is_string_dtype(s):
            txt = s.dropna().astype(str).str.strip()
            wanted = {str(x) for x in SENTINEL_NUMERIC} | set(SENTINEL_DATES)
            vc = txt.value_counts()
            for val, cnt in vc.items():
                if val in wanted:
                    found.append((val, int(cnt)))

        for val, cnt in found:
            matches = [
                k for k in keys
                if not (k.table == table and k.column == col)
                and k.n_distinct >= 3
                and abs(cnt - k.n_distinct) <= max(0.0, tol) * k.n_distinct
            ]
            exact = [k for k in matches if k.n_distinct == cnt]
            suspected = bool(matches) and cnt >= min_count
            if not suspected and (val in _SOFT_SENTINELS or cnt < min_count):
                continue          # 0 未命中就不報，否則整份報表都是 0；
                                  # 出現次數低於門檻的未命中值也不報，那是真值不是哨兵
            hits.append({
                "欄": col, "候選值": val, "出現次數": cnt,
                "占比": round(cnt / c["n_nonnull"], 6),
                "判定": "SUSPECTED_SENTINEL" if suspected else "未命中（僅記錄）",
                "命中方式": ("恰等於" if exact else "極接近") if suspected else None,
                "命中的group_by鍵": [
                    {"表": k.table, "欄": k.column, "相異值數": k.n_distinct}
                    for k in (exact or matches)
                ],
            })
            if suspected:
                c["suspected_sentinel"] = True
    return hits


# ══════════════════════════════════════════════════════════════════
# 逐欄的三桶判定（本腳本看得見的那幾條）
# ══════════════════════════════════════════════════════════════════
def judge_columns(df: pd.DataFrame, cols: list[dict[str, Any]], where: str,
                  bk: Buckets) -> None:
    names = [c["column"] for c in cols]

    for c in cols:
        w = f"{where}::{c['column']}"
        nr = c["null_rate"]

        # 缺失率（04 §二）
        if c["n_nonnull"] == 0:
            bk.add("info", "全空欄", w, "整欄無值",
                   "確認是上游沒給還是讀檔切錯；不進 staging，但要在契約留紀錄")
        elif nr >= NULL_DROP:
            bk.add("error", "缺失率", w, f"缺失率 {_pct(nr)} ≥ 80%",
                   "建議棄用。要保留必須在 contracts/<source>.yml 宣告處理方式，"
                   "否則補出來的是模型不是資料")
        elif nr >= NULL_ERROR:
            bk.add("error", "缺失率", w, f"缺失率 {_pct(nr)} ≥ 40%",
                   "必須在 contracts/<source>.yml 宣告處理方式（補值／標記／排除）"
                   "才能解除；>40–50% 時 MICE/迴歸補值補出來的是模型")
        elif nr >= NULL_WARN:
            bk.add("warning", "缺失率", w, f"缺失率 {_pct(nr)} ≥ 5%",
                   "寫進報告的「資料限制」節，並在 T4 缺失值總表填處理決議與理由")

        # Q9 helper column
        if c["helper_column"] and c["n_nonnull"]:
            extra, tail = _describe_helper(df, c["column"], names)
            bk.add("warning", "Q9", w,
                   f"欄名符合 ^Unnamed: \\d+$，缺失率 {_pct(c['null_rate'])}{extra}",
                   "標記 helper_column，保留但不當業務欄位 —— dropna(axis=1, how='all')"
                   " 不會刪它，於是它會被當成業務欄進入相關矩陣" + tail)

        # 04 §4.3 info
        if c["top10_shares"] and c["top10_shares"][0] > NEAR_CONSTANT:
            bk.add("info", "近乎常數欄", w,
                   f"單一值「{c['top10_values'][0]}」占 {_pct(c['top10_shares'][0])}",
                   "模型中無資訊量，M6 分群與 M7 迴歸可直接排除")
        if (c["n_unique"] > HIGH_CARDINALITY and not c["suspected_id"]
                and c["dtype_inferred"] == "object"):
            bk.add("info", "高基數類別欄", w, f"相異值 {c['n_unique']:,} 個且非 ID 欄",
                   "one-hot 會爆維度；改用頻率編碼、目標編碼或先做層級歸併")

        # 形狀量數（04 §二）
        if c["skewness"] is not None and abs(c["skewness"]) > SKEW_ABS:
            bk.add("info", "偏度", w, f"偏度 g1 = {c['skewness']:.3f}，|g1| > 1",
                   "標記「M3 需轉換」；直接跑迴歸違反常態假設")
        if c["cv"] is not None and c["cv"] > CV_HIGH:
            bk.add("info", "變異係數", w, f"CV = {c['cv']:.3f} > 1，分布極度分散",
                   "等寬分箱必失效，改用分位數（NTILE）分箱")

        # Q13 編碼序號缺口
        if c["top10_values"] and c["dtype_inferred"] == "object":
            _judge_code_prefix(df, c, w, bk)

        # Q10 全形半形混用 / 多餘空白
        if c["dtype_inferred"] == "object" and c["n_nonnull"]:
            t = df[c["column"]].dropna().astype(str)
            n_fw = int(t.str.contains(FULLWIDTH_RE, na=False).sum())
            n_sp = int((t != t.str.strip()).sum())
            if n_fw or n_sp:
                bk.add("warning", "Q10", w,
                       f"全形字元 {n_fw} 筆、前後多餘空白 {n_sp} 筆",
                       "groupby 前先做 NFKC 正規化與 strip，否則同一個值會被拆成兩組")

    # Q16 有金額欄但無幣別欄
    if any(MONEY_NAME_RE.search(n) for n in names) and \
            not any(CURRENCY_NAME_RE.search(n) for n in names):
        money = [n for n in names if MONEY_NAME_RE.search(n)][:3]
        bk.add("warning", "Q16", where,
               f"存在金額欄（{'、'.join(money)}）但全表無 currency 欄",
               "在契約明寫幣別假設，報告發 "
               "「WARNING: currency column absent, unit inferred as TWD "
               "from magnitude + locale」；多幣別未解析會讓 ROAS 差 30 倍")


def _describe_helper(df: pd.DataFrame, col: str, names: Sequence[str]) -> tuple[str, str]:
    """helper 欄到底是什麼？回傳 (事實補述, 行動補述)。

    三種一路放寬的判定，對應課程檔實測的三種樣貌：
      ① 與某欄逐列相同        → 是複本
      ② 值集合完全落在某欄內  → 是那一欄的一份「名單」（課程檔的 `Unnamed: 10`
         是 Step 5 那份 99 位客戶的 CAI 名單，因為 6687 被剔除而整體位移一列，
         逐列比對不相等，但值集合 99/99 落在 `客戶ID` 內）
      ③ 與某數值欄 |r| ≥ 0.95 → 進相關矩陣會踩中 18-E10
    """
    s = df[col]
    a = s.dropna()
    if a.empty:
        return "", ""
    for other in names:
        if other == col:
            continue
        try:
            if _norm_for_compare(a).tolist() == _norm_for_compare(df.loc[a.index, other]).tolist():
                return (f"，且逐列與「{other}」完全相同",
                        f"，並與「{other}」得到 r = 1.00（18-E10）")
        except Exception:  # noqa: BLE001
            continue
    va = set(_norm_for_compare(a).tolist())
    for other in names:
        if other == col:
            continue
        try:
            vb = set(_norm_for_compare(df[other].dropna()).tolist())
        except Exception:  # noqa: BLE001
            continue
        if va and va <= vb:
            return (f"，且 {len(va)} 個相異值全部落在「{other}」的值域內"
                    f"（{len(va)}/{len(vb)}）—— 這是一份名單，不是業務欄",
                    f"；它其實是「{other}」的一個子集名單，"
                    f"當成數值欄拿去算相關或分群完全沒有意義")
    if pd.api.types.is_numeric_dtype(s):
        best, r = None, 0.0
        for other in names:
            if other == col or not pd.api.types.is_numeric_dtype(df[other]):
                continue
            try:
                v = abs(float(s.corr(df[other])))
            except Exception:  # noqa: BLE001
                continue
            if np.isfinite(v) and v > r:
                best, r = other, v
        if best and r >= 0.95:
            return (f"，且與「{best}」的相關係數 |r| = {r:.4f}",
                    f"，並與「{best}」在相關矩陣得到 |r| = {r:.2f}（18-E10）")
    return "", ""


def _judge_code_prefix(df: pd.DataFrame, c: dict[str, Any], w: str, bk: Buckets) -> None:
    """Q13：類別編碼的數字前綴不連續，或不符主要命名慣例。"""
    vals = [str(v) for v in df[c["column"]].dropna().unique()]
    if not 2 <= len(vals) <= 200:
        return
    prefixes, seps, broken = [], [], []
    for v in vals:
        m = CODE_PREFIX_RE.match(v)
        if m:
            prefixes.append(int(m.group(1)))
            seps.append(m.group(2))
        else:
            broken.append(v)
    if len(prefixes) < max(3, 0.6 * len(vals)):
        return
    # 從 1 起算：編碼慣例是 01 開頭，最小值不是 1 就代表 01 那一類沒被匯出。
    # 實測課程檔的 15 種產業分類前綴是 [2,3,5,…,16]，缺的是 01 與 04 兩個。
    lo = min(1, min(prefixes))
    gaps = sorted(set(range(lo, max(prefixes) + 1)) - set(prefixes))
    main_sep = max(set(seps), key=seps.count)
    odd_sep = sorted({v for v, sp in zip(vals, seps) if sp != main_sep})
    if gaps or broken or odd_sep:
        parts = []
        if gaps:
            parts.append(f"數字前綴缺 {'、'.join(f'{g:02d}' for g in gaps[:8])}")
        if odd_sep:
            parts.append(f"分隔符破格：{'、'.join(odd_sep[:3])}（主要用「{main_sep}」）")
        if broken:
            parts.append(f"無數字前綴：{'、'.join(broken[:3])}")
        bk.add("warning", "Q13", w, "；".join(parts),
               "不要用 split('_') 解析編碼，會在破格那一列爆掉；"
               "改用正規表示式並對缺號向上游確認是否有未匯出的類別")


# ══════════════════════════════════════════════════════════════════
# 主流程
# ══════════════════════════════════════════════════════════════════
def profile_table(df: pd.DataFrame, source: str, table: str, bk: Buckets,
                  args: argparse.Namespace) -> dict[str, Any]:
    where = f"{source}::{table}"
    n_before = len(df)

    df, pivot = strip_pivot_noise(df, args.extra_noise_label,
                                  enabled=not args.no_pivot_clean)
    if pivot["已剔除列數"]:
        bk.add("error", "Q3", where,
               f"首欄「{pivot['首欄']}」有 {pivot['已剔除列數']} 列是樞紐雜訊列："
               f"{'、'.join(pivot['命中標籤'])}（原 {n_before} 列 → {len(df)} 列）",
               "本腳本已在剖析前剔除。下游讀這張表時必須做同樣的事 —— 不剔除的話，"
               "Grand Total 會變成一位「顧客」並衝上任何排行榜第一名。"
               "把剔除規則寫進 contracts/<source>.yml 與清理日誌")
    elif pivot["偵測到但未剔除"]:
        bk.add("error", "Q3", where,
               f"首欄「{pivot['首欄']}」有 {pivot['偵測到但未剔除']} 列是樞紐雜訊列："
               f"{'、'.join(pivot['命中標籤'])}，但 --no-pivot-clean 要求保留，"
               f"以下所有統計量都含這幾列",
               "確認這幾列真的是業務類別而非樞紐彙總；若是彙總，拿掉 "
               "--no-pivot-clean 重跑，本次的敘述統計全部作廢不可引用")

    if n_before in EXCEL_ROW_LIMITS:
        bk.add("error", "列上限截斷", where,
               f"列數恰為 {n_before:,}，等於 Excel 工作表上限",
               "這是上游被 Excel 截斷的指紋，不是巧合。回頭拿原始檔（實測案例："
               "Kaggle 原檔 1,296,675 列，被截成 1,048,575，遺失約 248,100 列）")

    cols = [profile_column(df, c, i, source, table, args.top)
            for i, c in enumerate(df.columns)]
    grain = infer_grain(df, cols, args.max_key_cols)

    if grain["整列重複列數"]:
        bk.add("error", "Q6", where,
               f"整列完全重複 {grain['整列重複組數']} 組、涉及 {grain['整列重複列數']} 列",
               "04 §三：只有「所有欄位皆同」才算真重複，這種要去重並記入清理日誌。"
               "不要放寬成「同鍵同日同金額」—— 實測那樣會誤刪 323 組真實交易")
    if grain["主鍵"] is None:
        bk.add("warning", "無主鍵", where, grain["粒度描述"],
               "在報告明寫「本表無主鍵」並列出最接近的複合鍵與重複率；"
               "粒度可能比你以為的更細（常見：以為是一筆交易，其實是一筆交易的一個品項）")
    else:
        for c in cols:
            if c["column"] in grain["主鍵"]:
                c["note"] = "；".join(x for x in (c["note"], "本表主鍵（粒度鍵）") if x)

    return {"表名": table, "列數": len(df), "原始列數": n_before,
            "欄數": df.shape[1], "粒度與主鍵": grain,
            "樞紐雜訊列": pivot, "欄位": cols, "哨兵值掃描": [],
            "_df": df}          # 暫存，哨兵掃描與三桶判定用完就刪，不進 JSON


def render_report(payload: dict[str, Any], bk: Buckets, brief: bool) -> str:
    L: list[str] = []
    W = 78
    L.append("=" * W)
    L.append("M1 資料剖析 — profile_dataset.py")
    L.append(f"專案：{payload['專案']}｜產生時間：{payload['產生時間']}")
    L.append("=" * W)

    for f in payload["檔案"]:
        inv = f["盤點"]
        L.append("")
        L.append("━" * W)
        L.append(f"檔案：{inv['檔名']}")
        L.append("━" * W)
        L.append(f"  · 路徑        {inv['路徑']}")
        L.append(f"  · 大小        {inv['MB']} MB"
                 + (f"（解壓後 {inv['解壓後大小MB']} MB）" if inv.get("解壓後大小MB") else ""))
        L.append(f"  · 格式        宣稱 .{inv['宣稱格式']}｜實際 {inv['實際格式']}"
                 f"｜magic {inv['magic']}")
        L.append(f"  · 表／分頁    {len(f['表'])} 個：{'、'.join(t['表名'] for t in f['表'])}")

        for t in f["表"]:
            g = t["粒度與主鍵"]
            L.append("")
            L.append("─" * W)
            L.append(f"表：{t['表名']}    {t['列數']:,} 列 × {t['欄數']} 欄"
                     + (f"（樞紐雜訊列已剔除 {t['樞紐雜訊列']['已剔除列數']} 列，"
                        f"原 {t['原始列數']:,} 列）" if t["樞紐雜訊列"]["已剔除列數"] else ""))
            L.append("─" * W)
            L.append(f"  粒度      {g['粒度描述']}")
            L.append(f"  主鍵      {'（無）' if not g['主鍵'] else ' × '.join(g['主鍵'])}"
                     f"{'  [' + g['主鍵種類'] + ']' if g['主鍵種類'] else ''}")
            if g["替代鍵"]:
                L.append(f"  替代鍵    {'、'.join(g['替代鍵'])}")
            L.append(f"  整列重複  {g['整列重複組數']} 組 / {g['整列重複列數']} 列")

            L.append("")
            L.append("  逐欄剖析")
            L.append("  " + _w("欄名", 22) + _w("推定型別", 18) + _w("缺失率", 10)
                     + _w("相異值", 10) + "標記")
            L.append("  " + "-" * (W - 2))
            for c in t["欄位"]:
                marks = []
                if c["suspected_id"]:
                    marks.append("suspected_id")
                if c["helper_column"]:
                    marks.append("helper")
                if c["suspected_sentinel"]:
                    marks.append("⛔SENTINEL")
                if c["practical_use"] not in marks:
                    marks.append(c["practical_use"])
                L.append("  " + _w(c["column"], 22) + _w(c["dtype_inferred"], 18)
                         + _w(_pct(c["null_rate"]), 10)
                         + _w(f"{c['n_unique']:,}", 10) + "｜".join(marks))

            nums = [c for c in t["欄位"] if c["mean"] is not None]
            if nums and not brief:
                L.append("")
                L.append("  數值欄敘述統計（min/q25/median/q75/max｜mean/std/p95/p99｜偏度/峰度/CV）")
                for c in nums:
                    L.append(f"    · {c['column']}")
                    L.append("        " + " / ".join(
                        _fmt(c[k]) for k in ("min", "q25", "median", "q75", "max")))
                    L.append("        " + " / ".join(
                        _fmt(c[k]) for k in ("mean", "std", "p95", "p99")))
                    L.append("        偏度 " + _fmt(c["skewness"]) + "｜峰度 "
                             + _fmt(c["kurtosis"]) + "｜CV " + _fmt(c["cv"])
                             + ("" if c["cv"] is not None else "（含 0 或負值，依 04 §二 不報）"))

            cats = [c for c in t["欄位"] if c["top10_values"]]
            if cats and not brief:
                L.append("")
                L.append(f"  類別欄 top-{payload['top_n']} 與占比"
                         "（接近逐列唯一的鍵欄不列）")
                for c in cats:
                    tail = c["tail_share"] or 0.0
                    L.append(f"    · {c['column']}（相異 {c['n_unique']:,}"
                             f"{'，尾部合併 ' + _pct(tail) if tail else ''}）")
                    for v, sh in zip(c["top10_values"], c["top10_shares"]):
                        L.append(f"        {_w(v, 34)} {_pct(sh):>8}")

            dts = [c for c in t["欄位"] if c["date_min"] is not None]
            if dts:
                L.append("")
                L.append("  日期欄")
                for c in dts:
                    L.append(f"    · {c['column']}  {c['date_min']} ~ {c['date_max']}"
                             f"｜相異日數 {c['n_distinct_dates']:,}"
                             f"｜缺日 {c['n_missing_dates']:,}")

            if t["哨兵值掃描"]:
                L.append("")
                L.append("  哨兵值掃描（候選集 -1/-9/-99/-999/0/9999/99999/999999 + 四個日期哨兵）")
                for h in t["哨兵值掃描"]:
                    tag = "⛔" if h["判定"] == "SUSPECTED_SENTINEL" else "·"
                    line = (f"    {tag} {h['欄']} = {h['候選值']}  出現 {h['出現次數']:,} 次"
                            f"（{_pct(h['占比'])}）→ {h['判定']}")
                    L.append(line)
                    for k in h["命中的group_by鍵"][:3]:
                        L.append(f"        └ {h['命中方式'] or '比對'} "
                                 f"{k['表']}.{k['欄']} 的相異值數 {k['相異值數']:,}")
            if t["樞紐雜訊列"]["命中標籤"]:
                L.append("")
                kept = t["樞紐雜訊列"]["偵測到但未剔除"]
                L.append("  Excel 樞紐雜訊列（"
                         + ("--no-pivot-clean：偵測到但保留，統計量含這幾列"
                            if kept else "已剔除，不是默默丟掉") + "）")
                for lab in t["樞紐雜訊列"]["命中標籤"]:
                    L.append(f"    ⛔ 首欄「{t['樞紐雜訊列']['首欄']}」= {lab}")

    L.append("")
    L.append("=" * W)
    L.append("三桶")
    L.append("=" * W)
    for sev, mark, title in (("error", "⛔", "error — 擋住，不准進 M2"),
                             ("warning", "⚠", "warning — 可往下，但必須進報告的「資料限制」節"),
                             ("info", "·", "info — 記錄，不阻擋")):
        items: list[Finding] = getattr(bk, sev)
        L.append("")
        L.append(f"{mark} {title}（{len(items)} 條）")
        L.append("-" * W)
        if not items:
            L.append("  （無）")
        for it in items:
            L.append(f"  {mark} [{it.rule}] {it.where}")
            L.append(f"      {it.fact}")
            L.append(f"      → {it.action}")
    L.append("")
    L.append("=" * W)
    code = payload["退出碼"]
    verdict = {0: "✅ 三桶無 error 無 warning → 可進 M2",
               1: "⛔ 有 error → 擋住，必須在 contracts/<source>.yml 宣告處理方式後重跑",
               2: "⚠ 只有 warning → 可進 M2，warning 條目必須進報告的「資料限制」節"}[code]
    L.append(f"結果：error {len(bk.error)}｜warning {len(bk.warning)}｜info {len(bk.info)}"
             f"　退出碼 {code}")
    L.append(verdict)
    L.append("=" * W)
    L.append("備註：本腳本只做 04 §一 的步驟 ①③④ 與 Q2/Q3/Q9/Q10/Q13/Q16 + 缺失率。")
    L.append("　　　Q1/Q5/Q6 完整版/Q7/Q8/Q11/Q12/Q14/Q15 需要契約、歷史載入或領域知識，")
    L.append("　　　由 check_schema_contract.py 與 check_data_quality.py 接手。")
    return "\n".join(L)


def collect_inputs(args: argparse.Namespace, p: Any, bk: Buckets) -> list[Path]:
    if args.files:
        out = []
        for f in args.files:
            fp = Path(f)
            if not fp.is_absolute():
                fp = (Path.cwd() / fp).resolve()
            if fp.exists():
                out.append(fp)
            else:
                bk.add("error", "檔案不存在", str(fp), "指定的路徑找不到檔案",
                       "確認路徑；相對路徑是相對於現在的工作目錄")
        return out
    exts = {".parquet", ".csv", ".tsv", ".xlsx", ".xlsm", ".xls"}
    return sorted(q for q in p.raw.rglob("*") if q.is_file() and q.suffix.lower() in exts)


def main() -> int:
    ap = argparse.ArgumentParser(
        description="M1 資料剖析：逐欄剖析 + 粒度主鍵判定 + 哨兵值與樞紐雜訊列掃描",
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("project", help="專案代號（路徑由 paths.project_dir 解析）")
    ap.add_argument("files", nargs="*", help="要剖析的檔案；省略則掃專案的 原始資料/")
    ap.add_argument("--from-db", action="store_true",
                    help="改為剖析專案 DuckDB 裡的所有表（走 db.connect）")
    ap.add_argument("--sheet", action="append", default=[],
                    help="只讀指定分頁（可重複）")
    ap.add_argument("--top", type=int, default=10, help="類別欄取前 N 名（預設 10）")
    ap.add_argument("--sentinel-tolerance", type=float, default=0.02,
                    help="哨兵值「極接近」的相對誤差，預設 0.02（2%%）")
    ap.add_argument("--sentinel-min-count", type=int, default=3,
                    help="哨兵值至少出現幾次才判定，預設 3")
    ap.add_argument("--group-key", action="append", type=int, default=[],
                    help="補一個外部已知的相異值數（例如另一份檔案的客戶數 100），可重複")
    ap.add_argument("--no-pivot-clean", action="store_true",
                    help="不要自動剔除 Grand Total／總計／(blank) 等樞紐雜訊列")
    ap.add_argument("--extra-noise-label", action="append", default=[],
                    help="額外的樞紐雜訊列標籤（可重複）")
    ap.add_argument("--max-key-cols", type=int, default=8,
                    help="複合鍵搜尋的候選欄上限，預設 8")
    ap.add_argument("--big-mb", type=int, default=200,
                    help="xlsx 解壓後超過幾 MB 就改走串流，預設 200")
    ap.add_argument("--stream-rows", type=int, default=200_000,
                    help="串流模式最多讀幾列，預設 200000")
    ap.add_argument("--report-to", choices=("stderr", "stdout", "none"), default="stderr",
                    help="人眼報表輸出到哪。預設 stderr，讓 stdout 保持乾淨的 JSON")
    ap.add_argument("--brief", action="store_true", help="報表精簡版（省略敘述統計與 top-N）")
    ap.add_argument("--no-write", action="store_true", help="不要寫落檔")
    args = ap.parse_args()

    bk = Buckets()
    p = project_dir(args.project, create=not args.no_write)

    files_payload: list[dict[str, Any]] = []

    if args.from_db:
        from db import connect  # 03 §7.1：唯一合法的連線介面，不准自己 duckdb.connect()
        with connect(args.project, read_only=True) as con:
            names = [r[0] for r in con.execute(
                "SELECT table_name FROM information_schema.tables "
                "WHERE table_schema = 'main' ORDER BY table_name").fetchall()]
            tables = {n: con.execute(f'SELECT * FROM "{n}"').df() for n in names}
        src = p.db.name
        entry: dict[str, Any] = {
            "盤點": {"路徑": str(p.db), "檔名": src, "MB": round(p.db.stat().st_size / 1e6, 2),
                     "宣稱格式": "duckdb", "實際格式": "duckdb", "magic": "（DuckDB 連線）",
                     "副檔名詐騙": False, "解壓後大小MB": None, "可整份載入": True},
            "表": [],
        }
        for name, df in tables.items():
            entry["表"].append(profile_table(df, src, name, bk, args))
        files_payload.append(entry)
    else:
        inputs = collect_inputs(args, p, bk)
        # 沒東西可剖析是 error（擋住），不是退出碼 3 —— 3 只留給腳本自己爆掉
        if not inputs and not bk.error:
            bk.add("error", "無輸入", args.project,
                   f"專案「{args.project}」的 {p.raw} 是空的，也沒有在命令列指定檔案",
                   f"把原始檔放進 {p.raw}，或直接指定："
                   f"python profile_dataset.py {args.project} <檔案>")
        for path in inputs:
            inv = inventory(path, bk, args.big_mb)
            entry = {"盤點": inv, "表": []}
            try:
                tables = read_tables(path, inv, bk, args.sheet or None, args.stream_rows)
            except Exception as e:  # noqa: BLE001
                bk.add("error", "讀檔失敗", path.name, f"{type(e).__name__}: {e}",
                       "確認檔案沒壞、格式受支援（parquet / csv / xlsx）；"
                       "xlsx 需要 openpyxl，parquet 需要 pyarrow")
                files_payload.append(entry)
                continue
            for name, df in tables.items():
                entry["表"].append(profile_table(df, path.name, name, bk, args))
            files_payload.append(entry)

    # ── 哨兵值掃描：group-by 鍵候選池取自**同一檔案的所有表** ──────
    # 必須等整個檔案的所有表都剖析完才做 —— 課程檔的 9999 在 `Step 2`，
    # 而「100」這個相異值數要從 `客戶ID` 欄取得，兩者可能不在同一張表。
    for entry in files_payload:
        keys: list[GroupKey] = []
        for t in entry["表"]:
            keys += collect_group_keys(t["表名"], t["欄位"])
        for extra in args.group_key:
            keys.append(GroupKey("（--group-key 外部指定）", f"n={extra}", extra))

        src = entry["盤點"]["檔名"]
        for t in entry["表"]:
            df = t.pop("_df")
            where = f"{src}::{t['表名']}"
            t["哨兵值掃描"] = scan_sentinels(
                df, t["欄位"], t["表名"], keys,
                args.sentinel_tolerance, args.sentinel_min_count)
            for h in t["哨兵值掃描"]:
                if h["判定"] != "SUSPECTED_SENTINEL":
                    continue
                k0 = h["命中的group_by鍵"][0]
                bk.add("error", "Q2", f"{where}::{h['欄']}",
                       f"值 {h['候選值']} 出現 {h['出現次數']:,} 次"
                       f"（占非空值 {_pct(h['占比'])}），"
                       f"{h['命中方式']}「{k0['表']}.{k0['欄']}」的相異值數"
                       f" {k0['相異值數']:,}",
                       "極可能是哨兵值不是真值。在 contracts/<source>.yml 的 "
                       "sentinels: 宣告處理方式（建議 to_null）才能解除。"
                       "實測：不排除會讓平均間隔從 10.79 變成 199.46（差 18.5 倍），"
                       "並把 CAI 整體壓向 0 —— 所有人都看起來『節奏穩定』，流失預警失效")
            judge_columns(df, t["欄位"], where, bk)
            del df

    payload: dict[str, Any] = {
        "腳本": "profile_dataset.py", "版本": VERSION,
        "產生時間": datetime.now().isoformat(timespec="seconds"),
        "專案": args.project, "top_n": args.top,
        "哨兵候選集": {"數值": list(SENTINEL_NUMERIC), "日期": list(SENTINEL_DATES)},
        "樞紐雜訊列標籤": list(PIVOT_NOISE_LABELS) + list(args.extra_noise_label),
        "檔案": files_payload,
        "三桶": {}, "退出碼": 0,
    }
    payload["三桶"] = {
        "error": [f.to_dict() for f in bk.error],
        "warning": [f.to_dict() for f in bk.warning],
        "info": [f.to_dict() for f in bk.info],
    }
    payload["退出碼"] = bk.gate()

    report = render_report(payload, bk, args.brief)

    if not args.no_write:
        try:
            rows = [c for f in files_payload for t in f["表"] for c in t["欄位"]]
            if rows:
                cols_order = [
                    "source", "table", "column", "ordinal",
                    "dtype_raw", "dtype_inferred",
                    "n_total", "n_nonnull", "null_rate", "n_unique", "unique_ratio",
                    "min", "q25", "median", "q75", "max", "mean", "std", "p95", "p99",
                    "skewness", "kurtosis", "cv",
                    "top10_values", "top10_shares", "tail_share",
                    "date_min", "date_max", "n_distinct_dates", "n_missing_dates",
                    "suspected_id", "helper_column", "suspected_sentinel",
                    "practical_use", "note",
                ]
                out = pd.DataFrame([{k: _jsonable(r.get(k)) for k in cols_order}
                                    for r in rows])
                out.to_csv(p.intake / "欄位總表.csv", index=False, encoding="utf-8-sig")
            (p.intake / "資料剖析.json").write_text(
                json.dumps(_jsonable(payload), ensure_ascii=False, indent=2),
                encoding="utf-8")
            (p.intake / "資料剖析報告.txt").write_text(report, encoding="utf-8")
        except OSError as e:
            print(f"⚠ 落檔失敗（{e}）— 剖析結果仍在 stdout，"
                  f"請確認 {p.intake} 可寫入", file=sys.stderr)

    if args.report_to == "stderr":
        print(report, file=sys.stderr)
    elif args.report_to == "stdout":
        print(report)

    print(json.dumps(_jsonable(payload), ensure_ascii=False, indent=2))
    return payload["退出碼"]


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except SystemExit:
        raise
    except Exception as exc:  # noqa: BLE001
        print(f"⛔ profile_dataset.py 本身失敗：{type(exc).__name__}: {exc}\n"
              f"   → 這是退出碼 3（腳本失敗），不是資料有問題。修腳本，不准手動略過。",
              file=sys.stderr)
        raise SystemExit(3) from exc
